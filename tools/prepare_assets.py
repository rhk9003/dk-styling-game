from __future__ import annotations

import json
import os
import re
from pathlib import Path

import openpyxl
from PIL import Image, ImageOps


WORKBOOK_NAME = "夏季主題小遊戲選鞋_商品資料.xlsx"
PHOTO_ROOT = r"\\10.10.1.252\電子商務部\04_商品及情境圖\商品照"
OUTPUT_DIR = Path("assets")
SHOE_DIR = OUTPUT_DIR / "shoes"
DATA_DIR = OUTPUT_DIR / "data"

CODE_RE = re.compile(r"\d{2}-\d{4}-\d{2}")
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

SEARCH_ROOTS = [
    "2026/2026春夏",
    "2025/2025秋冬/官網已上架",
    "2025/2025秋冬/官網已上架(已建檔)",
    "2025/2025春夏/官網已上架",
    "2023/2023春夏/官網已上架",
    "2025/原檔",
    "2024/原檔",
    "2023/原檔",
    "2024",
    "2023",
]


def norm_path(*parts: str) -> str:
    return os.path.join(PHOTO_ROOT, *parts)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " / ")).strip()


def short_name(name: str, code: str) -> str:
    value = re.sub(r"【[^】]+】", "", name)
    value = value.replace(code, "")
    return re.sub(r"\s+", " ", value).strip()


def color_from_name(name: str, code: str) -> str:
    tail = name.split(code, 1)[-1].strip() if code in name else ""
    return tail or "經典色"


def category_for(name: str) -> str:
    checks = [
        ("涼鞋", "涼鞋"),
        ("涼拖", "涼鞋"),
        ("穆勒", "穆勒鞋"),
        ("通勤", "通勤鞋"),
        ("戶外", "戶外鞋"),
        ("跑鞋", "戶外鞋"),
        ("老爹", "老爹鞋"),
        ("空氣", "空氣鞋"),
        ("休閒", "休閒鞋"),
        ("懶人", "懶人鞋"),
    ]
    for needle, label in checks:
        if needle in name:
            return label
    return "日常鞋"


def code_digits(code: str) -> str:
    return code.replace("-", "")


def find_candidate_dirs(code: str, allow_walk: bool = False) -> list[str]:
    matches: list[str] = []
    seen: set[str] = set()

    for root in SEARCH_ROOTS:
        abs_root = norm_path(*root.split("/"))
        direct = os.path.join(abs_root, code)
        if os.path.isdir(direct) and direct not in seen:
            matches.append(direct)
            seen.add(direct)

    if matches or not allow_walk:
        return matches

    for root in SEARCH_ROOTS:
        abs_root = norm_path(*root.split("/"))
        if not os.path.isdir(abs_root):
            continue
        for current, dirs, _files in os.walk(abs_root):
            for dirname in list(dirs):
                if dirname.startswith(code):
                    full = os.path.join(current, dirname)
                    if full not in seen:
                        matches.append(full)
                        seen.add(full)
            if matches and root in {"2024", "2023"}:
                break

    return matches


def dir_priority(path: str) -> tuple[int, int]:
    path_norm = path.replace("\\", "/")
    rules = [
        ("2026/2026春夏", 0),
        ("官網已上架", 1),
        ("2025/原檔", 2),
        ("2023/2023春夏", 3),
        ("2024/原檔", 4),
        ("2023/原檔", 5),
    ]
    for marker, score in rules:
        if marker in path_norm:
            return score, len(path_norm)
    return 9, len(path_norm)


def image_score(path: str, code: str) -> tuple[int, str]:
    name = Path(path).name.lower()
    stem = Path(path).stem.lower()
    digits = code_digits(code)

    score = 100
    if digits.lower() in stem:
        score -= 30
    if re.search(r"a0?1($|[_-])", stem):
        score -= 35
    elif re.search(r"a0?2($|[_-])", stem):
        score -= 25
    elif re.search(r"(^|[_-])0?1($|[_-])", stem):
        score -= 18
    if name.startswith("img_"):
        score -= 8
    if any(word in name for word in ["尺寸", "size", "banner", "desc", "detail", "注意"]):
        score += 40
    return score, name


def find_source_image(code: str) -> tuple[str | None, str | None]:
    dirs = sorted(find_candidate_dirs(code), key=dir_priority)
    if not dirs:
        dirs = sorted(find_candidate_dirs(code, allow_walk=True), key=dir_priority)
    for folder in dirs:
        images: list[str] = []
        for current, _dirs, files in os.walk(folder):
            for filename in files:
                if Path(filename).suffix.lower() in IMAGE_EXTENSIONS:
                    images.append(os.path.join(current, filename))
        if images:
            return sorted(images, key=lambda p: image_score(p, code))[0], folder
    return None, dirs[0] if dirs else None


def save_web_image(source: str, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode in {"RGBA", "LA"}:
            background = Image.new("RGB", image.size, "white")
            background.paste(image, mask=image.getchannel("A"))
            image = background
        else:
            image = image.convert("RGB")
        image.thumbnail((1100, 1100), Image.Resampling.LANCZOS)
        image.save(target, "WEBP", quality=84, method=6)


def main() -> None:
    workbook_path = Path(WORKBOOK_NAME)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    indexes = {header: idx for idx, header in enumerate(headers)}

    products = []
    missing: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[indexes["商品頁名稱"]])
        code_match = CODE_RE.search(name)
        if not code_match:
            continue

        code = code_match.group(0)
        image_source, source_dir = find_source_image(code)
        image_path = f"assets/shoes/{code}.webp"

        if image_source:
            save_web_image(image_source, Path(image_path))
        else:
            missing.append(code)

        products.append(
            {
                "id": str(row[indexes["商品頁序號"]]),
                "code": code,
                "url": f"https://www.dk-shoes.com.tw/SalePage/Index/{row[indexes['商品頁序號']]}",
                "name": name,
                "shortName": short_name(name, code),
                "color": color_from_name(name, code),
                "category": category_for(name),
                "price": int(row[indexes["售價"]] or 0),
                "sellingPoints": clean_text(row[indexes["銷售重點"]]),
                "features": [part.strip() for part in clean_text(row[indexes["商品特色"]]).split(" / ") if part.strip()],
                "image": image_path,
                "sourceImage": image_source,
                "sourceDir": source_dir,
            }
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "products.json").write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"products: {len(products)}")
    print(f"missing images: {', '.join(missing) if missing else 'none'}")
    print(f"output: {DATA_DIR / 'products.json'}")


if __name__ == "__main__":
    main()
